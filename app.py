import streamlit as str_lit
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font, Alignment
import io
import re
import os

# Sayfa Genişlik ve Başlık Ayarları
str_lit.set_page_config(page_title="LİSTECİ v2", layout="wide")

# CSS ile Arayüzü Modernleştirme ve Fare Tekerleği/Geniş Kutu Desteği
str_lit.markdown("""
    <style>
        .reportview-container { background: #f5f6fa; }
        .main .block-container { padding-top: 2rem; }
        h1 { color: #2c3e50; text-align: center; font-weight: bold; margin-bottom: 30px; }
        
        /* Ders Satırları ve Geniş Şube Kutuları */
        .ders-kart {
            background-color: white;
            padding: 15px;
            border-radius: 8px;
            border-left: 6px solid #e67e22;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
            margin-bottom: 15px;
        }
        /* Şube listelerinin dikeyde rahat kayması ve fare bilyası uyumu için */
        .sube-liste-alani {
            background-color: #fafafa;
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 10px;
            max-height: 140px;
            overflow-y: auto;
            font-size: 13px;
        }
        .sube-item {
            background-color: #dcdde1;
            padding: 4px 8px;
            margin: 4px 0;
            border-radius: 4px;
            display: flex;
            justify-content: space-between;
        }
    </style>
""", unsafe_allow_value=True)

# GÜVENLİK VE ŞİFRE EKRANI
def sifre_kontrol():
    if "sifreli_giris" not in str_lit.session_state:
        str_lit.session_state["sifreli_giris"] = False

    if not str_lit.session_state["sifreli_giris"]:
        str_lit.subheader("🔒 LİSTECİ - Yetkili Girişi")
        girilen_sifre = str_lit.text_input("Lütfen erişim şifresini giriniz:", type="password")
        # Şifreyi kendinize göre değiştirebilirsiniz (Örn: "asu123")
        if str_lit.button("Giriş Yap"):
            if girilen_sifre == "asu123":
                str_lit.session_state["sifreli_giris"] = True
                str_lit.rerun()
            else:
                str_lit.error("Hatalı şifre girdiniz!")
        return False
    return True

def temizle_metin(metin):
    m = str.maketrans("çğıöşüîâûÇĞİÖŞÜÎÂÛ", "CGIOSUUAUCGIOSUUAU")
    t = str(metin).translate(m).upper().strip()
    return re.sub(r'[^A-Z0-9]', '', t)

def main():
    str_lit.title("📋 LİSTECİ OTO-AVCI")
    
    # Hafıza havuzlarının tanımlanması
    if "dersler" not in str_lit.session_state: str_lit.session_state["dersler"] = []
    if "eslesmeler" not in str_lit.session_state: str_lit.session_state["eslesmeler"] = {}
    if "siralamalar" not in str_lit.session_state: str_lit.session_state["siralamalar"] = {}

    # 1. AŞAMA
    str_lit.markdown("### 1. AŞAMA: MATRİS DOSYASI")
    matris_file = str_lit.file_uploader("Sınav Görev Matrisini Seçin (.xlsx)", type=["xlsx"])
    
    # Şablonu doğrudan GitHub'dan otomatik okuyoruz (Kullanıcıya her seferinde şablon seçtirmemek için)
    sablon_yol = "sablon.xlsx"
    if not os.path.exists(sablon_yol):
        str_lit.error("⚠️ Depoda 'sablon.xlsx' dosyası bulunamadı! Lütfen GitHub'a şablonu yükleyin.")
        return

    if matris_file and str_lit.button("DERSLERİ OTOMATİK GETİR", type="primary"):
        wb = load_workbook(io.BytesIO(matris_file.read()), data_only=True)
        ws = wb.active
        
        y_sinir = ws.max_column
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=1, column=c).value or ws.cell(row=4, column=c).value or "").upper()
            if "GÖZETMENLİK" in val or "GOZETMEN" in val:
                y_sinir = c - 1
                break
                
        cols = range(5, y_sinir + 1)
        bulunan_dersler = []

        for c in cols:
            ders_val = ws.cell(row=4, column=c).value
            saat_val = ws.cell(row=3, column=c).value
            tarih_val = ws.cell(row=1, column=c).value
            
            if tarih_val is None:
                for merged in ws.merged_cells.ranges:
                    if ws.cell(row=1, column=c).coordinate in merged:
                        tarih_val = ws.cell(merged.min_row, merged.min_col).value
                        break
            try:
                t_str = tarih_val.strftime("%d.%m.%Y") if hasattr(tarih_val, 'strftime') else str(tarih_val).split(" ")[0]
                if "-" in t_str and len(t_str.split("-")[0]) == 4:
                    p = t_str.split("-"); t_str = f"{p[2]}.{p[1]}.{p[0]}"
            except:
                t_str = str(tarih_val)
            if t_str.lower() == "none": t_str = "Tarih_Yok"

            if ders_val is None:
                for merged in ws.merged_cells.ranges:
                    if ws.cell(row=4, column=c).coordinate in merged:
                        ders_val = ws.cell(merged.min_row, merged.min_col).value
                        break
            
            ders_kontrol_str = str(ders_val or "").upper()
            if "MUSIK" in ders_kontrol_str or "MUSİK" in ders_kontrol_str:
                if not ders_val: ders_val = "Dinî Musiki ve Nazariyatı"

            if ders_val and str(ders_val).strip().lower() != "none":
                bulunan_dersler.append({
                    "id": len(bulunan_dersler),
                    "tarih": t_str,
                    "saat": str(saat_val)[:5] if ":" in str(saat_val) else str(saat_val),
                    "ders_adi": str(ders_val).strip(),
                    "hucre": ws.cell(row=4, column=c).coordinate
                })
        
        str_lit.session_state["dersler"] = bulunan_dersler
        str_lit.session_state["eslesmeler"] = {d["id"]: [] for d in bulunan_dersler}
        str_lit.session_state["siralamalar"] = {d["id"]: "No" for d in bulunan_dersler}
        str_lit.success(f"Matristen {len(bulunan_dersler)} adet ders başarıyla çekildi!")

    # 2. AŞAMA
    if str_lit.session_state["dersler"]:
        str_lit.markdown("---")
        str_lit.markdown("### 2. AŞAMA: ÖĞRENCİ LİSTELERİNİ YÜKLE VE OTOMATİK EŞLEŞTİR")
        ogrenci_dosyalari = str_lit.file_uploader("Tüm şube/öğrenci listelerini buraya topluca bırakın:", type=["xlsx", "xls", "csv"], accept_multiple_files=True)
        
        if ogrenci_dosyalari and str_lit.button("🤖 OTOMATİK EŞLEŞTİRMEYİ BAŞLAT", type="secondary"):
            eslesen_sayisi = 0
            eslesmeyenler = []
            
            # Eşleşmeleri sıfırla
            for d in str_lit.session_state["dersler"]:
                str_lit.session_state["eslesmeler"][d["id"]] = []

            for f in ogrenci_dosyalari:
                dosya_adi_saf = os.path.splitext(f.name)[0]
                c_dosya = temizle_metin(dosya_adi_saf)
                is_haz_dosya = "HAZIRLIK" in c_dosya or "HZ" in c_dosya
                matched = False

                # Dosya içeriğini belleğe al
                f_bytes = f.read()

                for d in str_lit.session_state["dersler"]:
                    c_tam_ders = temizle_metin(d["ders_adi"])
                    is_haz_ders = "HAZIRLIK" in c_tam_ders or "HZ" in c_tam_ders
                    
                    if ("MUSIK" in c_dosya) and ("MUSIK" in c_tam_ders):
                        str_lit.session_state["eslesmeler"][d["id"]].append({"isim": f.name, "veri": f_bytes})
                        eslesen_sayisi += 1
                        matched = True
                        break

                    alt_dersler = [x.strip() for x in re.split(r'[-/\n]', d["ders_adi"]) if x.strip()]
                    for alt in alt_dersler:
                        c_ders = temizle_metin(alt)
                        if (c_dosya in c_ders or c_ders in c_dosya) or (is_haz_ders and is_haz_dosya):
                            str_lit.session_state["eslesmeler"][d["id"]].append({"isim": f.name, "veri": f_bytes})
                            eslesen_sayisi += 1
                            matched = True
                            break
                    if matched: break
                
                if not matched:
                    eslesmeyenler.append(f.name)

            if eslesmeyenler:
                str_lit.warning(f"{eslesen_sayisi} dosya eşleşti. Eşleşmeyen listeler mevcut:\n" + "\n".join(eslesmeyenler))
            else:
                str_lit.success(f"Mükemmel! Yüklenen {eslesen_sayisi} listenin tamamı derslerle eşleşti!")

        # OLUŞTURULAN SINAV LİSTELERİ PANELİ
        str_lit.markdown("---")
        str_lit.markdown("### 📂 OLUŞTURULAN SINAV LİSTESİ (KARMA VE TEKLİ)")
        
        for d in str_lit.session_state["dersler"]:
            d_id = d["id"]
            eslesen_listeler = str_lit.session_state["eslesmeler"].get(d_id, [])
            
            # Web arayüzünde her dersi şık bir kartta gösteriyoruz
            str_lit.markdown(f"""
            <div class="ders-kart">
                <strong>[{d['tarih']} | {d['saat']}]</strong> - {d['ders_adi']} (Hücre: {d['hucre']})
            </div>
            """, unsafe_allow_html=True)
            
            col_sol, col_sag = str_lit.columns([3, 1])
            
            with col_sol:
                # Fare bilyası uyumlu yüksek kutu simülasyonu
                if eslesen_listeler:
                    sube_html = '<div class="sube-liste-alani">'
                    for h_list in eslesen_listeler:
                        sube_html += f'<div class="sube-item"><span>📄 {h_list["isim"]}</span></div>'
                    sube_html += '</div>'
                    str_lit.markdown(sube_html, unsafe_allow_html=True)
                else:
                    str_lit.info("Bu ders için henüz bir şube listesi eşleşmedi.")
            
            with col_sag:
                str_lit.session_state["siralamalar"][d_id] = str_lit.radio(
                    "Sıralama:", ["No", "Ad", "Soyad"], 
                    key=f"sort_{d_id}", horizontal=True
                )
        
        # 3. AŞAMA: TANZİM VE İNDİRME
        str_lit.markdown("---")
        str_lit.markdown("### 📥 3. AŞAMA: LİSTELERİ TANZİM ET VE ZIP OLARAK AL")
        
        if str_lit.button("🔥 TÜM LİSTELERİ HAZIRLA VE İNDİRME LİNKİ OLUŞTUR", type="primary"):
            import zipfile
            zip_buffer = io.BytesIO()
            m_maketrans = str.maketrans("çğıöşüîâûÇĞİÖŞÜÎÂÛ", "CGIOSUUAUCGIOSUUAU")
            
            # Matrisi tekrar okuyoruz
            matris_file.seek(0)
            ws_g = load_workbook(io.BytesIO(matris_file.read()), data_only=True).active
            
            basarili_sayisi = 0
            
            with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                for d in str_lit.session_state["dersler"]:
                    d_id = d["id"]
                    dosyalar = str_lit.session_state["eslesmeler"].get(d_id, [])
                    if not dosyalar: continue
                    
                    havuz = []
                    c_tam_ders = str(d["ders_adi"]).translate(m_maketrans).upper()
                    ders_hazirlik_mi = "HAZIRLIK" in c_tam_ders or "HZ" in c_tam_ders
                    
                    for f_data in dosyalar:
                        try:
                            c_dosya = str(f_data["isim"]).translate(m_maketrans).upper()
                            if ders_hazirlik_mi or "HAZIRLIK" in c_dosya or "HZ" in c_dosya:
                                t = pd.read_excel(io.BytesIO(f_data["veri"]), skiprows=2).iloc[:, [1, 2, 3]].copy()
                                t.columns = ['No', 'Ad', 'Soyad']
                                t['Sorumlu'] = "HAZIRLIK KOORDİNATÖRLÜĞÜ"
                            else:
                                df_r = pd.read_excel(io.BytesIO(f_data["veri"]))
                                cols_to_take = min(4, len(df_r.columns))
                                t = df_r.iloc[:, list(range(cols_to_take))].copy()
                                if cols_to_take == 3:
                                    t.columns = ['No', 'Ad', 'Soyad']
                                    t['Sorumlu'] = "Belirtilmedi"
                                else:
                                    t.columns = ['No', 'Ad', 'Soyad', 'Sorumlu']
                                    t['Sorumlu'] = t['Sorumlu'].ffill().fillna("Belirtilmedi")
                            havuz.append(t)
                        except:
                            continue
                    
                    if not havuz: continue
                    df = pd.concat(havuz, ignore_index=True)
                    sk = str_lit.session_state["siralamalar"].get(d_id, "No")
                    
                    if sk == "No": 
                        df = df.sort_values(by=['Sorumlu', 'No'])
                    elif sk == "Ad":
                        df['sk'] = df['Ad'].apply(lambda x: str(x).translate(m_maketrans).upper().strip())
                        df = df.sort_values(by=['Sorumlu', 'sk', 'No']).drop(columns=['sk'])
                    else:
                        df['sk'] = df['Soyad'].apply(lambda x: str(x).translate(m_maketrans).upper().strip())
                        df = df.sort_values(by=['Sorumlu', 'sk', 'No']).drop(columns=['sk'])

                    c_idx = column_index_from_string("".join(filter(str.isalpha, d["hucre"])))
                    gorevler = []
                    for r in range(6, 400):
                        target = ws_g.cell(row=r, column=4)
                        color = str(target.fill.start_color.index) if target.fill else ""
                        txt = str(target.value or "").upper()
                        if ("FFFF0000" in color or "FF0000" in color) or ("SINAVA GİRECEK ÖĞRENCİ" in txt): break
                        s_val = str(ws_g.cell(row=r, column=c_idx).value or "").strip()
                        if s_val and s_val.lower() != "none" and len(s_val) < 10:
                            gorevler.append({'sinif': s_val, 'goz': str(target.value or "")})
                            
                    if not gorevler: continue
                    
                    plan = []; kalan = len(df)
                    k_sin = [g for g in gorevler if g['sinif'] not in ['205', '305']]
                    b_sin = [g for g in gorevler if g['sinif'] in ['205', '305']]
                    
                    for g in k_sin:
                        if kalan <= 0: break
                        m = min(34, kalan); plan.append({'b': g, 'm': m}); kalan -= m
                    if b_sin and kalan > 0:
                        p = kalan // len(b_sin); a = kalan % len(b_sin)
                        for g in b_sin:
                            m = p + (1 if a > 0 else 0); plan.append({'b': g, 'm': m}); kalan -= m; a = max(0, a-1)

                    f_wb = load_workbook(sablon_yol)
                    idx = 0
                    for p_od od in plan:
                        g, kap = p_od['b'], p_od['m']
                        dilim = df.iloc[idx : idx + kap]
                        target_sheet_name = "Sayfa2" if g['sinif'] in ['205', '305'] else "Sayfa1"
                        if target_sheet_name not in f_wb.sheetnames:
                            target_sheet_name = f_wb.sheetnames[0]
                            
                        sh = f_wb.copy_worksheet(f_wb[target_sheet_name])
                        sh.title = f"Sinif_{g['sinif']}"
                        sh['B3'], sh['G5'], sh['C4'], sh['G4'] = d["ders_adi"], d["saat"], g['sinif'], d["tarih"]
                        
                        hocalar = dilim['Sorumlu'].unique()
                        hoca_metni = "\n".join([str(h).strip() for h in hocalar if str(h).strip() and str(h).lower() != 'nan'])
                        r_alt = '75' if g['sinif'] in ['205', '305'] else '45'
                        sh[f'E{r_alt}'] = g['goz']
                        t_h = sh[f'A{r_alt}']
                        t_h.value = hoca_metni
                        t_h.alignment = Alignment(wrapText=True, vertical='center')
                        t_h.font = Font(bold=True, size=10 if len(hocalar) < 3 else 8)
                        
                        for i, (_, row) in enumerate(dilim.iterrows()):
                            sh[f'C{7+i}'], sh[f'D{7+i}'], sh[f'E{7+i}'] = row['No'], str(row['Ad']).upper(), str(row['Soyad']).upper()
                            sh[f'F{7+i}'].value = str(row['Sorumlu']).upper()
                            sh[f'F{7+i}'].font = Font(size=10); sh[f'F{7+i}'].alignment = Alignment(shrink_to_fit=True)
                        idx += kap
                        
                    for s in ["Sayfa1", "Sayfa2"]: 
                        if s in f_wb.sheetnames and len(f_wb.sheetnames) > 1: 
                            f_wb.remove(f_wb[s])
                            
                    temiz_ders_adi = re.sub(r'[^\w]', '_', d["ders_adi"][:30])
                    dosya_adi = f"{d['tarih'].replace('.','_')}_{d['saat'].replace(':','.')}_{temiz_ders_adi}.xlsx"
                    
                    # Dosyayı hafızada kaydedip ZIP içine ekliyoruz
                    b_out = io.BytesIO()
                    f_wb.save(b_out)
                    klasor_adi = d["tarih"].replace(".", "_")
                    zip_file.writestr(f"{klasor_adi}/{dosya_adi}", b_out.getvalue())
                    basarili_sayisi += 1
            
            if basarili_sayisi > 0:
                str_lit.success(f"🎉 {basarili_sayisi} ders için listeler paketlendi!")
                str_lit.download_button(
                    label="📥 TÜM LİSTELERİ ZIP OLARAK BİLGİSAYARINA İNDİR",
                    data=zip_buffer.getvalue(),
                    file_name="Tanzim_Edilen_Sinav_Listeleri.zip",
                    mime="application/zip"
                )
            else:
                str_lit.error("Eşleşen şube listesi bulunamadığı için dosya üretilemedi.")

if __name__ == "__main__":
    if sifre_kontrol():
        main()
